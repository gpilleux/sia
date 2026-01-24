"""
XLSX File Reader - Microsoft Excel Spreadsheet Parser

Extracts text from XLSX files including:
- All worksheets (sheets)
- Cell values in row-major order
- Formatted output with sheet names

Domain: Skills (Infrastructure)
Bounded Context: File Processing
Pattern: Strategy (concrete implementation)

Dependencies:
    - openpyxl: Pure Python library for Excel files
    
Invariant:
    XlsxReader.read(valid_xlsx) → non_empty_text
    ∧ XlsxReader.read(corrupted_xlsx) → CorruptedFileError
    ∧ XlsxReader ∈ AbstractFileReader.registry["xlsx"]

REQ-011: File Reader Skills System
QUANT-011-003: Concrete Readers Implementation
"""

from pathlib import Path
from typing import TYPE_CHECKING

from .base import AbstractFileReader, CorruptedFileError, validate_file_exists

if TYPE_CHECKING:
    from openpyxl import Workbook


class XlsxReader(AbstractFileReader):
    """
    Extract text from XLSX (Microsoft Excel) files.
    
    This reader uses openpyxl to extract text from all worksheets in an
    Excel file, preserving the sheet structure and cell layout.
    
    Features:
        - Read-only mode for memory efficiency
        - Data-only mode (evaluates formulas to values)
        - Tab-separated cell values per row
        - Sheet names preserved as section headers
        - Empty rows skipped
    
    Implementation Notes:
        - Uses read_only=True for large file support
        - Uses data_only=True to get calculated values, not formulas
        - iter_rows() with values_only=True for performance
        - Workbook properly closed after reading
    
    Edge Cases Handled:
        - Corrupted ZIP archives (InvalidFileException)
        - Password-protected files (detected and rejected)
        - Empty sheets (skipped)
        - Cells with formulas (values extracted, not formulas)
        - None values (converted to empty string)
    
    Example:
        >>> reader = XlsxReader()
        >>> text = reader.read(Path("spreadsheet.xlsx"))
        >>> print(text[:200])
    """
    
    @classmethod
    def get_extension(cls) -> str:
        """Return supported extension: 'xlsx'"""
        return "xlsx"
    
    def read(self, filepath: Path) -> str:
        """
        Extract all text from an XLSX file.
        
        Args:
            filepath: Path to XLSX file
            
        Returns:
            Extracted text with sheet sections and tab-separated values
            
        Raises:
            FileNotFoundError: If file doesn't exist
            CorruptedFileError: If file is corrupted, invalid, or password-protected
        """
        validate_file_exists(filepath)
        
        # Lazy import to avoid forcing dependency
        try:
            from openpyxl import load_workbook
            from openpyxl.utils.exceptions import InvalidFileException
        except ImportError as e:
            raise ImportError(
                "openpyxl not installed. "
                "Use: uv run --with openpyxl python your_script.py"
            ) from e
        
        # Open XLSX file in read-only mode
        try:
            workbook = load_workbook(
                str(filepath),
                read_only=True,   # Memory-efficient streaming mode
                data_only=True    # Get formula values, not formulas
            )
        except InvalidFileException as e:
            raise CorruptedFileError(
                f"Invalid XLSX structure - file may be corrupted: {e}"
            ) from e
        except Exception as e:
            # Detect password-protected files
            error_msg = str(e).lower()
            if "password" in error_msg or "encrypted" in error_msg:
                raise CorruptedFileError(
                    "Password-protected XLSX files are not supported"
                ) from e
            # Generic error handling
            raise CorruptedFileError(
                f"Failed to open XLSX file: {e}"
            ) from e
        
        try:
            text_parts = []
            
            # Process all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Add sheet header
                text_parts.append(f"\n=== SHEET: {sheet_name} ===\n")
                
                # Extract rows
                sheet_content = self._extract_sheet_text(sheet)
                if sheet_content:
                    text_parts.append(sheet_content)
            
            return "\n".join(text_parts)
        
        finally:
            # Always close workbook to free resources
            workbook.close()
    
    def _extract_sheet_text(self, sheet) -> str:
        """
        Extract text from a worksheet.
        
        Args:
            sheet: openpyxl Worksheet object
            
        Returns:
            Tab-separated rows of text
        """
        rows_text = []
        
        # Use iter_rows with values_only for performance
        for row in sheet.iter_rows(values_only=True):
            # Convert all cells to strings, handling None values
            row_values = []
            for cell in row:
                if cell is not None:
                    # Convert to string, preserving numbers and dates
                    row_values.append(str(cell))
                else:
                    row_values.append("")
            
            # Join with tabs and skip completely empty rows
            row_text = "\t".join(row_values)
            if row_text.strip():  # Skip rows with only whitespace
                rows_text.append(row_text)
        
        return "\n".join(rows_text)
